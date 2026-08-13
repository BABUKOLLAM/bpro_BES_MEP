import base64
import csv
import io
from datetime import datetime

from pytz import timezone as pytz_timezone, UTC

from odoo import fields, models
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class BproAttendanceImport(models.TransientModel):
    _name = "bpro.attendance.import"
    _description = "Attendance Log Import"

    attachment = fields.Binary(required=True, attachment=False)
    filename = fields.Char()
    result_summary = fields.Text(readonly=True, copy=False)

    def _parse_rows(self):
        """Expected columns (header row required, case-insensitive):
        badge_id, date (YYYY-MM-DD), check_in (HH:MM), check_out
        (HH:MM, optional) - the daily-summary shape virtually every
        punch-device's bundled management software can export, one row
        per employee per day. Not tied to any specific device brand -
        ME Polymers' actual device wasn't known at build time (2026-08-13
        scoping decision), so this is the device-agnostic path; a live
        connector is a separate future addition once that's confirmed.
        """
        if not self.filename:
            raise UserError("Choose a file first.")
        raw = base64.b64decode(self.attachment)
        ext = self.filename.rsplit(".", 1)[-1].lower()
        rows = []
        if ext == "csv":
            text = raw.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append({(k or "").strip().lower(): (v or "").strip() if isinstance(v, str) else v
                             for k, v in row.items()})
        elif ext in ("xlsx", "xlsm"):
            if not load_workbook:
                raise UserError("XLSX support is unavailable on this server - export as .csv instead.")
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            sheet = wb.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1))
            header = [str(c.value or "").strip().lower() for c in header_row]
            for excel_row in sheet.iter_rows(min_row=2):
                values = [c.value for c in excel_row]
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append(dict(zip(header, values)))
        else:
            raise UserError("Unsupported file type - upload a .csv or .xlsx export.")
        return rows

    @staticmethod
    def _to_utc_datetime(date_raw, time_raw, tz_name):
        """Punch-device exports report local time; hr.attendance stores
        UTC. Localize against the employee's own tz (falling back to
        Asia/Kolkata - every client on this deployment is India-based)
        before converting, or worked_hours/overtime would be off by the
        UTC offset for every imported record."""
        if not time_raw:
            return False
        tz = pytz_timezone(tz_name or "Asia/Kolkata")
        if isinstance(date_raw, datetime):
            date_part = date_raw.date()
        else:
            date_part = datetime.strptime(str(date_raw).strip(), "%Y-%m-%d").date()
        if isinstance(time_raw, datetime):
            time_part = time_raw.time()
        else:
            time_part = datetime.strptime(str(time_raw).strip(), "%H:%M").time()
        localized = tz.localize(datetime.combine(date_part, time_part))
        return localized.astimezone(UTC).replace(tzinfo=None)

    def action_import(self):
        self.ensure_one()
        rows = self._parse_rows()
        Employee = self.env["hr.employee"].sudo()
        Attendance = self.env["hr.attendance"].sudo()
        imported = 0
        skipped = []
        for i, row in enumerate(rows, start=2):
            badge = str(row.get("badge_id") or "").strip()
            date_raw = row.get("date")
            if not badge or not date_raw:
                skipped.append(f"Row {i}: missing badge_id or date")
                continue
            employee = Employee.search([("barcode", "=", badge)], limit=1)
            if not employee:
                skipped.append(f"Row {i}: no employee with Badge ID '{badge}'")
                continue
            try:
                check_in = self._to_utc_datetime(date_raw, row.get("check_in"), employee.tz)
                check_out = self._to_utc_datetime(date_raw, row.get("check_out"), employee.tz)
            except ValueError as e:
                skipped.append(f"Row {i}: could not parse date/time - {e}")
                continue
            if not check_in:
                skipped.append(f"Row {i}: missing check_in time")
                continue
            if check_out and check_out <= check_in:
                skipped.append(f"Row {i}: check_out is not after check_in for {employee.name}")
                continue
            existing = Attendance.search([
                ("employee_id", "=", employee.id),
                ("check_in", "=", check_in),
            ], limit=1)
            if existing:
                skipped.append(f"Row {i}: attendance already recorded for {employee.name} at {check_in} UTC")
                continue
            Attendance.create({
                "employee_id": employee.id,
                "check_in": check_in,
                "check_out": check_out,
            })
            imported += 1
        self.result_summary = (
            f"Imported {imported} attendance record(s) out of {len(rows)} row(s).\n\n"
            + (f"Skipped {len(skipped)}:\n" + "\n".join(skipped) if skipped else "No rows skipped.")
        )
        return {
            "type": "ir.actions.act_window",
            "res_model": "bpro.attendance.import",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
